"""Low-level parsing of B3 "Movimentação" exports, CSV or XLSX.

B3 offers the same report in both formats, with the same columns; which one you
get depends on which button you press. The spreadsheet is normalised into the
text the CSV would have carried and both then follow one code path — see
:func:`parse_xlsx`, and :func:`_cell_text` for why the conversion goes through
``Decimal`` rather than ``str``.

The export looks like::

    Entrada/Saída,Data,Movimentação,Produto,Instituição,Quantidade,Preço unitário,Valor da Operação
    Credito,15/03/2024,Rendimento,XPTO11 - EXEMPLO FDO INV IMOB,CORRETORA EXEMPLO CCTVM S/A.,100," R$ 0,10 "," R$ 10,00 "

Quirks handled here:

* pt-BR numbers (``1.234,56``) and currency prefixes (`` R$ 0,07 ``);
* ``-`` used as "not applicable" for price/amount;
* products in four shapes: ``TICKER - Company``, ``CDB - CODE - Bank``,
  ``Futuro - WING21``, ``Opção de Compra - PETRM59 - PETR`` and bare names
  such as ``Tesouro Renda+ Aposentadoria Extra 2065``;
* the same broker written nine different ways;
* UTF-8 with or without BOM, plus latin-1 fallback for older exports.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from app.domain.enums import AssetKind

# Canonical header names (accent/casing insensitive lookup keys).
COL_DIRECTION = "entradasaida"
COL_DATE = "data"
COL_MOVEMENT = "movimentacao"
COL_PRODUCT = "produto"
COL_INSTITUTION = "instituicao"
COL_QUANTITY = "quantidade"
COL_UNIT_PRICE = "precounitario"
COL_TOTAL = "valordaoperacao"

REQUIRED_COLUMNS = (COL_DIRECTION, COL_DATE, COL_MOVEMENT, COL_PRODUCT)

_NA_TOKENS = {"", "-", "--", "n/a", "na", "nulo"}
_TICKER_RE = re.compile(r"^[A-Z0-9]{4}[0-9]{1,2}[A-Z]?$")
_OPTION_RE = re.compile(r"^(op[cç][aã]o de (compra|venda))\s*-\s*(?P<code>[A-Z0-9]+)", re.IGNORECASE)
_FUTURE_RE = re.compile(r"^futuros?\s*-\s*(?P<code>[A-Z0-9]+)", re.IGNORECASE)
_FIXED_INCOME_PREFIXES = ("CDB", "LCI", "LCA", "LC", "LF", "CRI", "CRA", "DEB", "RDB", "LIG")


#: An ``.xlsx`` is a zip archive, so the file says what it is without being
#: asked — the same mechanical test the upload endpoint uses for ``%PDF-``.
XLSX_MAGIC = b"PK\x03\x04"
#: How far down the sheet to look for the header row. B3's spreadsheet has it
#: first, but an export that has been opened and re-saved sometimes gains a
#: title line above it, and finding the header is cheaper than being wrong.
XLSX_HEADER_SEARCH_ROWS = 10


class CsvFormatError(ValueError):
    """Raised when the uploaded file is not a recognisable B3 export."""


def is_xlsx(payload: bytes) -> bool:
    """True when the payload is a spreadsheet rather than delimited text."""
    return bytes(payload[:4]) == XLSX_MAGIC


def strip_accents(value: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c))


def normalize_key(value: str) -> str:
    """Header/label normaliser: accent-free, lowercase, alphanumeric only."""
    return re.sub(r"[^a-z0-9]", "", strip_accents(value or "").lower())


def decode_bytes(payload: bytes) -> str:
    """Decode an upload, tolerating BOMs and legacy latin-1 exports."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def sniff_delimiter(sample: str) -> str:
    """B3 exports use ``,`` but some regional builds emit ``;``."""
    header = sample.splitlines()[0] if sample.splitlines() else ""
    return ";" if header.count(";") > header.count(",") else ","


def parse_decimal(value: str | None) -> Decimal | None:
    """Parse ``" R$ 1.234,56 "`` / ``"0,17"`` / ``"-"`` into a Decimal."""
    if value is None:
        return None
    text = strip_accents(str(value)).strip()
    if text.lower() in _NA_TOKENS:
        return None
    text = re.sub(r"(?i)(r\$|brl|us\$|usd|\$)", "", text).strip()
    if not text or text in _NA_TOKENS:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").strip()
    if "," in text:  # pt-BR: dot is the thousands separator
        text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in ("", "-", "."):
        return None
    try:
        parsed = Decimal(text)
    except InvalidOperation:
        return None
    return -parsed if negative else parsed


def parse_date(value: str) -> date:
    """Parse the ``Data`` column (dd/mm/yyyy, with ISO as a fallback)."""
    text = (value or "").strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise CsvFormatError(f"unrecognised date: {value!r}")


def normalize_broker(name: str) -> str:
    """Collapse the many spellings of one broker into a canonical name.

    The sample export contains nine institution strings for four brokers
    (e.g. ``XP INVESTIMENTOS CCTVM S/A``, ``XP INVESTIMENTOS CCTVM S/A.`` and
    ``XP INVESTIMENTOS CORRETORA DE CAMBIO...``). Normalising keeps broker
    filters usable and makes de-duplication robust across exports.
    """
    raw = re.sub(r"\s+", " ", (name or "").strip()).rstrip(".")
    if not raw:
        return "Desconhecida"
    key = normalize_key(raw)
    known = (
        ("clear", "Clear Corretora"),
        ("rico", "Rico Investimentos"),
        ("nuinvest", "Nu Invest"),
        ("nuinvestimentos", "Nu Invest"),
        ("nufinanceira", "Nu Invest"),
        ("xpinvestimentos", "XP Investimentos"),
        ("modal", "Modal DTVM"),
        ("inter", "Banco Inter"),
        ("btg", "BTG Pactual"),
        ("itau", "Itaú Corretora"),
        ("bradesco", "Bradesco Corretora"),
        ("nikos", "Nikos DTVM"),
        ("genial", "Genial Investimentos"),
        ("avenue", "Avenue"),
        ("toro", "Toro Investimentos"),
    )
    for needle, canonical in known:
        if key.startswith(needle) or needle in key:
            return canonical
    # Fallback: title-case the first three words ("BANCO X DTVM LTDA" -> "Banco X Dtvm")
    return " ".join(w.capitalize() for w in raw.split()[:3])


def classify_asset_kind(ticker: str, product: str) -> AssetKind:
    """Infer the instrument family from the ticker suffix / product text."""
    upper_product = strip_accents(product or "").upper()
    ticker = (ticker or "").upper()

    if upper_product.startswith("TESOURO") or "TESOURO" in upper_product:
        return AssetKind.TREASURY
    if _FUTURE_RE.match(product or "") or upper_product.startswith("FUTURO"):
        return AssetKind.FUTURE
    if _OPTION_RE.match(product or "") or upper_product.startswith("OPCAO"):
        return AssetKind.OPTION
    head = upper_product.split(" - ")[0].strip()
    if head in _FIXED_INCOME_PREFIXES:
        return AssetKind.FIXED_INCOME

    match = re.match(r"^([A-Z]{4})(\d{1,2})$", ticker)
    if match:
        suffix = int(match.group(2))
        if suffix == 11:
            # 11 is shared by FIIs, ETFs and units — disambiguate on the name.
            # ETF markers win: an index fund is still described as a "FUNDO".
            if any(k in upper_product for k in ("INDICE", "ISHARES", "ETF", "INDEX", "TRACKER")):
                return AssetKind.ETF
            if any(k in upper_product for k in ("FDO", "FUNDO", "FII", "IMOB", "FIAGRO", "RECEBIVEIS")):
                return AssetKind.FII
            # What is left is a unit (TAEE11, ALUP11): a bundle of shares of the
            # same company. It trades, pays and is taxed like a share, so it is
            # one — a separate class would only fragment the allocation view.
            return AssetKind.STOCK
        if suffix in (1, 2, 9, 10, 12, 13, 14, 15, 16):
            # Subscription rights (1/2) and receipts (9..16) are transient.
            return AssetKind.SUBSCRIPTION
        if suffix in (3, 4, 5, 6, 7, 8):
            if any(k in upper_product for k in ("FDO", "FUNDO", "FII")):
                return AssetKind.FII
            return AssetKind.STOCK
        if suffix in (31, 32, 33, 34, 35, 39):
            return AssetKind.BDR
    return AssetKind.OTHER


@dataclass(slots=True)
class ParsedProduct:
    """Result of splitting the ``Produto`` column."""

    ticker: str
    name: str
    kind: AssetKind


def parse_product(product: str) -> ParsedProduct:
    """Split ``Produto`` into a stable ticker plus a human readable name."""
    raw = re.sub(r"\s+", " ", (product or "").strip())
    if not raw:
        return ParsedProduct(ticker="DESCONHECIDO", name="Desconhecido", kind=AssetKind.OTHER)

    option = _OPTION_RE.match(raw)
    if option:
        code = option.group("code").upper()
        return ParsedProduct(ticker=code, name=raw, kind=AssetKind.OPTION)

    future = _FUTURE_RE.match(raw)
    if future:
        code = future.group("code").upper()
        return ParsedProduct(ticker=code, name=raw, kind=AssetKind.FUTURE)

    parts = [p.strip() for p in raw.split(" - ") if p.strip()]
    head = parts[0].upper() if parts else ""

    # Fixed income: "CDB - CDB7246C5YO - BANCO GUANABARA S/A"
    if head in _FIXED_INCOME_PREFIXES and len(parts) >= 2:
        code = parts[1].upper()
        issuer = " - ".join(parts[2:]) if len(parts) > 2 else ""
        name = f"{head} {issuer}".strip() if issuer else raw
        return ParsedProduct(ticker=code, name=name, kind=AssetKind.FIXED_INCOME)

    # Standard equities/FIIs/ETFs: "PETR4 - PETROLEO BRASILEIRO S.A."
    if len(parts) >= 2 and _TICKER_RE.match(head):
        name = " - ".join(parts[1:])
        return ParsedProduct(ticker=head, name=name, kind=classify_asset_kind(head, raw))

    if _TICKER_RE.match(head) and len(parts) == 1:
        return ParsedProduct(ticker=head, name=raw, kind=classify_asset_kind(head, raw))

    # Bare descriptions (Tesouro Direto and friends): synthesise a ticker.
    synthetic = re.sub(r"[^A-Z0-9]+", "-", strip_accents(raw).upper()).strip("-")[:40]
    return ParsedProduct(ticker=synthetic, name=raw, kind=classify_asset_kind("", raw))


@dataclass(slots=True)
class RawRow:
    """One CSV line, normalised but not yet interpreted."""

    line_number: int
    direction_raw: str
    trade_date: date
    movement: str
    product_raw: str
    institution_raw: str
    quantity: Decimal
    unit_price: Decimal | None
    total: Decimal | None
    product: ParsedProduct = field(init=False)
    broker: str = field(init=False)

    def __post_init__(self) -> None:
        self.product = parse_product(self.product_raw)
        self.broker = normalize_broker(self.institution_raw)


@dataclass(slots=True)
class ParseResult:
    rows: list[RawRow]
    errors: list[dict]
    total_lines: int


def _header_map(fieldnames: list[str]) -> dict[str, str]:
    return {normalize_key(name): name for name in fieldnames if name}


def parse_csv(payload: bytes | str) -> ParseResult:
    """Parse a full export into :class:`RawRow` objects.

    Bad lines never abort the import: they are collected in ``errors`` and
    surfaced in the import log so nothing fails silently.
    """
    text = decode_bytes(payload) if isinstance(payload, (bytes, bytearray)) else payload
    text = text.lstrip("﻿")
    if not text.strip():
        raise CsvFormatError("the file is empty")

    delimiter = sniff_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise CsvFormatError("the file has no header row")
    return _build_rows(list(reader.fieldnames), reader, first_data_line=2)


def _build_rows(
    fieldnames: list[str], records: Iterable[dict], first_data_line: int
) -> ParseResult:
    """Turn header-keyed records into :class:`RawRow`s.

    Shared by both formats: every rule about what a movement means lives here,
    so a spreadsheet and a CSV of the same export cannot drift apart.
    """
    headers = _header_map(fieldnames)
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise CsvFormatError(
            "unexpected layout — missing column(s): "
            + ", ".join(missing)
            + f". Found: {', '.join(str(name) for name in fieldnames if name)}"
        )

    def cell(row: dict, key: str) -> str:
        source = headers.get(key)
        return (row.get(source) or "").strip() if source else ""

    rows: list[RawRow] = []
    errors: list[dict] = []
    total = 0

    for index, row in enumerate(records, start=first_data_line):
        if not any((v or "").strip() for v in row.values()):
            continue
        total += 1
        try:
            direction_raw = cell(row, COL_DIRECTION)
            movement = re.sub(r"\s+", " ", cell(row, COL_MOVEMENT))
            product_raw = cell(row, COL_PRODUCT)
            if not movement and not product_raw:
                continue
            quantity = parse_decimal(cell(row, COL_QUANTITY)) or Decimal(0)
            rows.append(
                RawRow(
                    line_number=index,
                    direction_raw=direction_raw,
                    trade_date=parse_date(cell(row, COL_DATE)),
                    movement=movement,
                    product_raw=product_raw,
                    institution_raw=cell(row, COL_INSTITUTION),
                    quantity=quantity,
                    unit_price=parse_decimal(cell(row, COL_UNIT_PRICE)),
                    total=parse_decimal(cell(row, COL_TOTAL)),
                )
            )
        except Exception as exc:  # noqa: BLE001 — one bad row must not kill the import
            errors.append({"line": index, "error": str(exc), "raw": {k: v for k, v in row.items() if v}})

    return ParseResult(rows=rows, errors=errors, total_lines=total)


def _cell_text(value: object) -> str:
    """Render a spreadsheet cell as the text the CSV would have held.

    Normalising to text rather than teaching the row builder about types keeps
    one set of rules for both formats — but the conversion has to be careful in
    two places:

    * A number reaches us as a ``float``, and ``str()`` on a small one yields
      ``'1e-05'``, which the pt-BR number parser strips to nonsense. Going via
      ``Decimal`` and formatting with ``'f'`` keeps plain decimal notation.
      ``Decimal(str(x))`` — not ``Decimal(x)`` — because the shortest
      round-tripping repr is the number the sheet displays, whereas the exact
      binary value of ``0.07`` is ``0.070000000000000006661338147750939242541...``.
    * A date arrives as ``datetime``, and the parser expects ``dd/mm/yyyy``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):  # a stray checkbox is not a quantity
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    return str(value).strip()


def _find_header(rows: list[tuple], limit: int) -> int:
    """Index of the row that carries the column names, or -1."""
    for index, row in enumerate(rows[:limit]):
        names = _header_map([_cell_text(cell) for cell in row])
        if all(column in names for column in REQUIRED_COLUMNS):
            return index
    return -1


def parse_xlsx(payload: bytes) -> ParseResult:
    """Parse the spreadsheet form of the B3 export.

    Read-only and ``data_only``: the file is a report, so formulas are of no
    interest and the cached results are what the user sees on screen.
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover — packaging error
        raise CsvFormatError(
            "reading .xlsx files needs the openpyxl package"
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — any zip/XML failure is a bad upload
        raise CsvFormatError(f"could not read the spreadsheet: {exc}") from exc

    try:
        if not workbook.worksheets:
            raise CsvFormatError("the spreadsheet has no sheets")
        sheet = workbook.worksheets[0]
        # B3's file declares ``<dimension ref="A1"/>`` — one cell — and then
        # writes the whole report underneath it. Read-only mode believes the
        # declaration and stops after the first cell, so the export arrives as
        # a lone "Entrada/Saída" and nothing else. Recomputing the extent from
        # the rows actually present is the documented way out, and it costs
        # nothing on a file whose header is honest.
        sheet.reset_dimensions()
        grid = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not grid:
        raise CsvFormatError("the file is empty")

    header_index = _find_header(grid, XLSX_HEADER_SEARCH_ROWS)
    if header_index < 0:
        found = ", ".join(_cell_text(cell) for cell in grid[0] if _cell_text(cell))
        raise CsvFormatError(
            "unexpected spreadsheet layout — no header row with the B3 columns "
            f"in the first {XLSX_HEADER_SEARCH_ROWS} rows. First row: {found}"
        )

    fieldnames = [_cell_text(cell) for cell in grid[header_index]]
    records = [
        dict(zip(fieldnames, (_cell_text(cell) for cell in row)))
        for row in grid[header_index + 1 :]
    ]
    # Spreadsheet rows are 1-based and the header is one of them.
    return _build_rows(fieldnames, records, first_data_line=header_index + 2)


def parse_movements(payload: bytes | str) -> ParseResult:
    """Parse a B3 export in whichever of its two shapes it arrived.

    The file decides, by its own magic bytes; nothing asks the user to say
    which button they pressed on B3's site.
    """
    if isinstance(payload, (bytes, bytearray)) and is_xlsx(payload):
        return parse_xlsx(bytes(payload))
    return parse_csv(payload)
